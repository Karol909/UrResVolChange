import pandas as pd
import openpyxl
from datetime import datetime
import os


class ExcelLog:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Data"


        date = datetime.now()
        self.file_path = os.path.join("Logs", date.strftime("%d-%b-%Y_%H-%M-%S") + "_log.xlsx")

        self.wb.save(filename=self.file_path)
    
    def save(self, data_number, data_type, data):

        df = pd.DataFrame([[data_number, data_type, data]],
                          columns=["Data number", "Data type", "Data"])

     
        try:
            book = openpyxl.load_workbook(self.file_path)
            if self.ws.title not in book.sheetnames:
                book.create_sheet(self.ws.title)
                book.save(self.file_path)
            with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = writer.sheets[self.ws.title]
                startrow = sheet.max_row

                write_header = startrow == 1
                df.to_excel(writer, sheet_name=self.ws.title, index=False, header=write_header, startrow=startrow)
        except FileNotFoundError:
            df.to_excel(self.file_path, index=False)
    

    def save_row(self, table_row):
        table_row = [float(x) for x in table_row]
        df = pd.DataFrame([table_row],
                        columns=["Target", "Max volume","First reading", "Steps remaining", "Step delay", "Time","Iterations","Times jiggled","Last reading","Error"])

        try:
            book = openpyxl.load_workbook(self.file_path)
            if self.ws.title not in book.sheetnames:
                book.create_sheet(self.ws.title)
                book.save(self.file_path)
            with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = writer.sheets[self.ws.title]

                startcol = 7  # column H (0-indexed)
                
                # Find first empty row in this column
                startrow = 1
                while sheet.cell(row=startrow+1, column=startcol+1).value is not None:
                    startrow += 1

                # Only write header if first time
                write_header = startrow == 1

                df.to_excel(writer, sheet_name=self.ws.title, index=False, header=write_header,
                            startrow=startrow, startcol=startcol)
        except FileNotFoundError:
            df.to_excel(self.file_path, index=False, startcol=startcol)


    def save_excel(self):
        """Force save workbook to disk (if open in memory)."""
        try:
            self.wb.save(self.file_path)
            print(f"[ExcelLog] Workbook saved: {self.file_path}")
        except Exception as e:
            print(f"[ExcelLog] Error while saving workbook: {e}")

    def unload(self):
        """Unload workbook and release memory."""
        try:
            if hasattr(self, "wb") and self.wb:
                self.wb.close()
                self.wb = None
                self.ws = None
                print("[ExcelLog] Workbook unloaded from memory.")
        except Exception as e:
            print(f"[ExcelLog] Error while unloading: {e}")